from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Any, cast
from xml.sax.saxutils import escape

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.core.supabase_client import supabase


REPORT_TYPES = {
    "executive",
    "client",
    "dataset",
    "analysis",
    "investigation",
}

EXPORT_FORMATS = {"pdf", "xlsx", "csv", "json"}


class ReportService:
    """
    Internal-audit analytics reporting service.

    The report structure is aligned to the key reporting elements described
    by ICAI SIA 370 (Reporting Results), but the platform does not claim that
    the underlying audit engagement was conducted in accordance with the SIAs.
    """

    @staticmethod
    def _fetch_all() -> dict[str, list[dict[str, Any]]]:
        clients_response = (
            supabase.table("clients")
            .select("*")
            .order("created_at", desc=True)
            .execute()
        )
        datasets_response = (
            supabase.table("datasets")
            .select("*")
            .order("upload_date", desc=True)
            .execute()
        )
        analyses_response = (
            supabase.table("analyses")
            .select("*")
            .order("created_at", desc=True)
            .execute()
        )
        findings_response = (
            supabase.table("findings")
            .select("*")
            .order("created_at", desc=True)
            .execute()
        )

        return {
            "clients": cast(
                list[dict[str, Any]],
                clients_response.data or [],
            ),
            "datasets": cast(
                list[dict[str, Any]],
                datasets_response.data or [],
            ),
            "analyses": cast(
                list[dict[str, Any]],
                analyses_response.data or [],
            ),
            "findings": cast(
                list[dict[str, Any]],
                findings_response.data or [],
            ),
        }

    @staticmethod
    def _latest_completed_by_dataset(
        analyses: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        """
        Resolve one CURRENT completed assessment per dataset.

        The analyses table remains the execution/audit trail. Auditor-facing
        reporting uses the latest completed result for each dataset unless an
        explicit analysis_id is supplied by an existing internal route.
        """
        latest: dict[str, dict[str, Any]] = {}

        for item in sorted(
            analyses,
            key=lambda row: str(
                row.get("created_at") or ""
            ),
            reverse=True,
        ):
            if str(
                item.get("analysis_status") or ""
            ).upper() != "COMPLETED":
                continue

            dataset_id = str(
                item.get("dataset_id") or ""
            )

            if (
                dataset_id
                and dataset_id not in latest
            ):
                latest[dataset_id] = item

        return list(latest.values())

    @staticmethod
    def get_context() -> dict[str, Any]:
        data = ReportService._fetch_all()
        current_analyses = ReportService._latest_completed_by_dataset(data["analyses"])

        client_map = {
            str(item.get("id")): item
            for item in data["clients"]
            if item.get("id")
        }
        dataset_map = {
            str(item.get("id")): item
            for item in data["datasets"]
            if item.get("id")
        }

        clients = [
            {
                "id": item.get("id"),
                "client_name": item.get("client_name"),
                "client_code": item.get("client_code"),
                "industry": item.get("industry"),
                "client_status": item.get("client_status"),
            }
            for item in data["clients"]
        ]

        datasets: list[dict[str, Any]] = []
        for item in data["datasets"]:
            client = client_map.get(str(item.get("client_id")))
            datasets.append(
                {
                    "id": item.get("id"),
                    "client_id": item.get("client_id"),
                    "dataset_name": item.get("dataset_name"),
                    "file_type": item.get("file_type"),
                    "total_records": item.get("total_records"),
                    "total_columns": item.get("total_columns"),
                    "upload_status": item.get("upload_status"),
                    "upload_date": item.get("upload_date"),
                    "client_name": (
                        client.get("client_name")
                        if client
                        else None
                    ),
                }
            )

        analyses: list[dict[str, Any]] = []
        for item in current_analyses:
            dataset = dataset_map.get(str(item.get("dataset_id")))
            client = client_map.get(str(item.get("client_id")))
            analyses.append(
                {
                    "id": item.get("id"),
                    "client_id": item.get("client_id"),
                    "dataset_id": item.get("dataset_id"),
                    "analysis_status": item.get("analysis_status"),
                    "created_at": item.get("created_at"),
                    "total_transactions": item.get("total_transactions"),
                    "high_risk_count": item.get("high_risk_count"),
                    "medium_risk_count": item.get("medium_risk_count"),
                    "low_risk_count": item.get("low_risk_count"),
                    "dataset_name": (
                        dataset.get("dataset_name")
                        if dataset
                        else None
                    ),
                    "client_name": (
                        client.get("client_name")
                        if client
                        else None
                    ),
                }
            )

        return {
            "clients": clients,
            "datasets": datasets,
            "analyses": analyses,
            "report_types": sorted(REPORT_TYPES),
            "export_formats": sorted(EXPORT_FORMATS),
        }

    @staticmethod
    def _validate_scope(
        report_type: str,
        client_id: str | None,
        dataset_id: str | None,
        analysis_id: str | None,
    ) -> None:
        if report_type not in REPORT_TYPES:
            raise HTTPException(
                status_code=400,
                detail="Invalid report type.",
            )

        if report_type == "client" and not client_id:
            raise HTTPException(
                status_code=400,
                detail="Client is required.",
            )

        if report_type == "dataset" and not dataset_id:
            raise HTTPException(
                status_code=400,
                detail="Dataset is required.",
            )

        if report_type in {"analysis", "investigation"} and not analysis_id:
            raise HTTPException(
                status_code=400,
                detail="Analysis is required.",
            )

    @staticmethod
    def build_report(
        report_type: str,
        client_id: str | None = None,
        dataset_id: str | None = None,
        analysis_id: str | None = None,
        risk_level: str | None = None,
        detection_source: str | None = None,
    ) -> dict[str, Any]:
        report_type = report_type.strip().lower()

        ReportService._validate_scope(
            report_type,
            client_id,
            dataset_id,
            analysis_id,
        )

        data = ReportService._fetch_all()
        current_analyses = ReportService._latest_completed_by_dataset(data["analyses"])

        client_map = {
            str(item.get("id")): item
            for item in data["clients"]
            if item.get("id")
        }
        dataset_map = {
            str(item.get("id")): item
            for item in data["datasets"]
            if item.get("id")
        }
        analysis_map = {
            str(item.get("id")): item
            for item in data["analyses"]
            if item.get("id")
        }

        selected_analyses = list(current_analyses)

        if client_id:
            selected_analyses = [
                item
                for item in selected_analyses
                if str(item.get("client_id")) == str(client_id)
            ]

        if dataset_id:
            selected_analyses = [
                item
                for item in selected_analyses
                if str(item.get("dataset_id")) == str(dataset_id)
            ]

        if analysis_id:
            # Backward compatibility for existing internal links/routes.
            selected_analyses = [
                item
                for item in data["analyses"]
                if str(item.get("id")) == str(analysis_id)
            ]

        if not selected_analyses:
            raise HTTPException(
                status_code=404,
                detail="No completed analysis data found for the selected context.",
            )

        selected_analysis_ids = {
            str(item.get("id"))
            for item in selected_analyses
            if item.get("id")
        }

        selected_findings = [
            item
            for item in data["findings"]
            if str(item.get("analysis_id")) in selected_analysis_ids
        ]

        if risk_level and risk_level != "All":
            selected_findings = [
                item
                for item in selected_findings
                if str(item.get("risk_level")) == risk_level
            ]

        # Kept for backward route compatibility. The revised UI does not expose
        # this technical filter to non-technical users.
        if detection_source and detection_source != "All":
            selected_findings = [
                item
                for item in selected_findings
                if detection_source
                in ReportService._as_string_list(
                    item.get("detection_sources")
                )
            ]

        selected_dataset_ids = {
            str(item.get("dataset_id"))
            for item in selected_analyses
            if item.get("dataset_id")
        }
        selected_client_ids = {
            str(item.get("client_id"))
            for item in selected_analyses
            if item.get("client_id")
        }

        selected_datasets = [
            dataset_map[item_id]
            for item_id in selected_dataset_ids
            if item_id in dataset_map
        ]
        selected_clients = [
            client_map[item_id]
            for item_id in selected_client_ids
            if item_id in client_map
        ]

        risk_counts = Counter(
            str(item.get("risk_level") or "Unknown")
            for item in selected_findings
        )

        audit_check_counter: Counter[str] = Counter()
        relationship_indicator_count = 0
        unusual_behaviour_count = 0

        enriched_findings: list[dict[str, Any]] = []

        for item in selected_findings:
            triggered_rules = ReportService._as_string_list(
                item.get("triggered_rules")
            )
            anomaly_reasons = ReportService._as_string_list(
                item.get("anomaly_reasons")
            )
            network_reasons = ReportService._network_reasons(item)

            audit_check_counter.update(triggered_rules)

            if network_reasons:
                relationship_indicator_count += 1

            if anomaly_reasons:
                unusual_behaviour_count += 1

            analysis = analysis_map.get(str(item.get("analysis_id")))
            dataset = (
                dataset_map.get(str(analysis.get("dataset_id")))
                if analysis
                else None
            )
            client = (
                client_map.get(str(analysis.get("client_id")))
                if analysis
                else None
            )

            reasons = ReportService._reason_list(item.get("reasons"))
            primary_observation = (
                triggered_rules[0]
                if triggered_rules
                else (
                    reasons[0]
                    if reasons
                    else "Transaction requires auditor review"
                )
            )

            enriched_findings.append(
                {
                    "id": item.get("id"),
                    "transaction_id": item.get("transaction_id"),
                    "risk_score": ReportService._number(
                        item.get("risk_score")
                    ),
                    "risk_level": item.get("risk_level"),
                    "primary_observation": primary_observation,
                    "reasons": reasons,
                    "triggered_rules": triggered_rules,
                    "anomaly_reasons": anomaly_reasons,
                    "network_reasons": network_reasons,
                    "analysis_id": item.get("analysis_id"),
                    "dataset_id": dataset.get("id") if dataset else None,
                    "dataset_name": (
                        dataset.get("dataset_name")
                        if dataset
                        else None
                    ),
                    "client_id": client.get("id") if client else None,
                    "client_name": (
                        client.get("client_name")
                        if client
                        else None
                    ),
                    "created_at": item.get("created_at"),
                    # Technical evidence is preserved for traceability but moved
                    # to the appendix/export detail rather than executive sections.
                    "technical_evidence": {
                        "rule_score": ReportService._number(
                            item.get("rule_score")
                        ),
                        "anomaly_score": ReportService._number(
                            item.get("anomaly_score")
                        ),
                        "network_score": ReportService._number(
                            item.get("network_score")
                        ),
                        "detection_sources": (
                            ReportService._as_string_list(
                                item.get("detection_sources")
                            )
                        ),
                    },
                }
            )

        enriched_findings.sort(
            key=lambda item: (
                ReportService._risk_rank(
                    str(item.get("risk_level") or "")
                ),
                float(item.get("risk_score") or 0),
            ),
            reverse=True,
        )

        total_findings = len(enriched_findings)
        total_transactions = sum(
            int(item.get("total_transactions") or 0)
            for item in selected_analyses
        )
        high_critical = (
            risk_counts.get("High", 0)
            + risk_counts.get("Critical", 0)
        )
        review_rate = (
            round(
                (total_findings / total_transactions) * 100,
                2,
            )
            if total_transactions > 0
            else 0
        )

        audit_observations = [
            {
                "observation": name,
                "count": count,
            }
            for name, count in audit_check_counter.most_common()
        ]

        recommendations = ReportService._build_recommendations(
            risk_counts=risk_counts,
            audit_observations=audit_observations,
            relationship_indicator_count=relationship_indicator_count,
            unusual_behaviour_count=unusual_behaviour_count,
        )

        executive_summary = (
            ReportService._build_executive_summary(
                total_transactions=total_transactions,
                total_findings=total_findings,
                high_critical=high_critical,
                critical=risk_counts.get("Critical", 0),
                selected_clients=selected_clients,
                selected_datasets=selected_datasets,
            )
        )

        scope_text = ReportService._scope_text(
            report_type=report_type,
            selected_clients=selected_clients,
            selected_datasets=selected_datasets,
            selected_analyses=selected_analyses,
        )

        return {
            "report_type": report_type,
            "report_title": ReportService._report_title(report_type),
            "report_status": "DRAFT - FOR AUDITOR REVIEW",
            "generated_at": (
                datetime.now(timezone.utc)
                .isoformat(timespec="seconds")
                .replace("+00:00", "Z")
            ),
            "scope": {
                "client_id": client_id,
                "dataset_id": dataset_id,
                "analysis_id": analysis_id,
                "risk_level": risk_level or "All",
                "detection_source": detection_source or "All",
            },
            "clients": selected_clients,
            "datasets": selected_datasets,
            "analyses": selected_analyses,
            "engagement": {
                "objective": (
                    "To support risk-based review of uploaded ERP/financial "
                    "transaction data and identify transactions requiring "
                    "further auditor attention."
                ),
                "scope": scope_text,
                "approach": (
                    "The platform reviewed the selected transaction population "
                    "using configured audit checks, unusual-behaviour analysis "
                    "and vendor-relationship indicators, and consolidated the "
                    "results into transaction-level risk findings."
                ),
            },
            "executive_summary": executive_summary,
            "summary": {
                "total_clients": len(selected_clients),
                "total_datasets": len(selected_datasets),
                "total_analyses": len(selected_analyses),
                "total_transactions": total_transactions,
                "total_findings": total_findings,
                "transactions_requiring_review": total_findings,
                "review_rate": review_rate,
                "high_critical_findings": high_critical,
                "relationship_indicator_findings": relationship_indicator_count,
                "unusual_behaviour_findings": unusual_behaviour_count,
                "risk_counts": {
                    "Low": risk_counts.get("Low", 0),
                    "Medium": risk_counts.get("Medium", 0),
                    "High": risk_counts.get("High", 0),
                    "Critical": risk_counts.get("Critical", 0),
                },
            },
            "audit_observations": audit_observations,
            # Retain old key for compatibility with any other existing caller.
            "top_triggered_rules": [
                {
                    "rule": item["observation"],
                    "count": item["count"],
                }
                for item in audit_observations[:10]
            ],
            "recommendations": recommendations,
            "findings": enriched_findings,
            "methodology_note": (
                "This is a technology-assisted internal audit analytics report. "
                "It supports auditor review and does not constitute a statutory "
                "audit opinion, certification, or independent assurance report."
            ),
            "standards_note": (
                "The report layout follows the reporting principles of ICAI "
                "SIA 370 by presenting objectives/scope/approach, an executive "
                "summary, significant observations and recommended corrective "
                "actions. The platform does not assert that the underlying "
                "engagement itself was conducted in accordance with the SIAs."
            ),
        }

    @staticmethod
    def export_report(
        report: dict[str, Any],
        export_format: str,
    ) -> tuple[bytes, str, str]:
        export_format = export_format.lower()

        if export_format not in EXPORT_FORMATS:
            raise HTTPException(
                status_code=400,
                detail="Invalid export format.",
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = (
            f"internal_audit_{report['report_type']}_"
            f"{timestamp}"
        )

        if export_format == "pdf":
            return (
                ReportService._to_pdf(report),
                "application/pdf",
                f"{base_name}.pdf",
            )

        if export_format == "xlsx":
            return (
                ReportService._to_xlsx(report),
                (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                f"{base_name}.xlsx",
            )

        if export_format == "csv":
            return (
                ReportService._to_csv(report),
                "text/csv; charset=utf-8",
                f"{base_name}.csv",
            )

        return (
            json.dumps(
                report,
                indent=2,
                default=str,
            ).encode("utf-8"),
            "application/json",
            f"{base_name}.json",
        )

    @staticmethod
    def _to_pdf(report: dict[str, Any]) -> bytes:
        buffer = BytesIO()

        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=16 * mm,
            leftMargin=16 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
            title=report["report_title"],
            author="AI Audit Risk Analysis Platform",
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            alignment=TA_CENTER,
            fontSize=20,
            leading=25,
            textColor=colors.HexColor("#3B1D76"),
            spaceAfter=8,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["BodyText"],
            alignment=TA_CENTER,
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#64748B"),
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=12,
            leading=15,
            textColor=colors.HexColor("#3B1D76"),
            spaceBefore=10,
            spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["BodyText"],
            fontSize=8.5,
            leading=12,
            alignment=TA_LEFT,
        )
        small_style = ParagraphStyle(
            "Small",
            parent=styles["BodyText"],
            fontSize=7,
            leading=9,
        )

        client_name = ReportService._primary_client_name(report)
        dataset_name = ReportService._primary_dataset_name(report)

        story: list[Any] = [
            Spacer(1, 22 * mm),
            Paragraph(
                escape(report["report_title"]),
                title_style,
            ),
            Paragraph(
                escape(report.get("report_status") or "DRAFT"),
                subtitle_style,
            ),
            Spacer(1, 8 * mm),
        ]

        cover_rows = [
            ["Client", client_name],
            ["Dataset / Scope", dataset_name],
            [
                "Generated",
                ReportService._format_date(
                    report.get("generated_at")
                ),
            ],
            ["Prepared By", "AI Audit Risk Analysis Platform"],
        ]
        cover_table = Table(
            cover_rows,
            colWidths=[42 * mm, 116 * mm],
        )
        cover_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (0, -1),
                        colors.HexColor("#F1ECFA"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (0, -1),
                        colors.HexColor("#3B1D76"),
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (0, -1),
                        "Helvetica-Bold",
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.4,
                        colors.HexColor("#D8DDE8"),
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        8.5,
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                ]
            )
        )
        story.extend(
            [
                cover_table,
                Spacer(1, 9 * mm),
                Paragraph(
                    escape(report["methodology_note"]),
                    small_style,
                ),
                PageBreak(),
            ]
        )

        story.append(
            Paragraph(
                "1. Engagement Overview",
                section_style,
            )
        )
        engagement = report.get("engagement") or {}
        overview_rows = [
            ["Objective", engagement.get("objective") or "-"],
            ["Scope", engagement.get("scope") or "-"],
            ["Approach", engagement.get("approach") or "-"],
        ]
        overview_table = Table(
            [
                [
                    row[0],
                    Paragraph(
                        escape(str(row[1])),
                        body_style,
                    ),
                ]
                for row in overview_rows
            ],
            colWidths=[32 * mm, 126 * mm],
        )
        overview_table.setStyle(
            ReportService._pdf_detail_table_style()
        )
        story.extend([overview_table, Spacer(1, 4 * mm)])

        story.append(
            Paragraph(
                "2. Executive Summary",
                section_style,
            )
        )
        story.append(
            Paragraph(
                escape(report.get("executive_summary") or "-"),
                body_style,
            )
        )
        story.append(Spacer(1, 3 * mm))

        summary = report["summary"]
        risk = summary["risk_counts"]

        summary_rows = [
            ["Transactions Reviewed", summary["total_transactions"]],
            [
                "Transactions Requiring Review",
                summary["transactions_requiring_review"],
            ],
            ["Review Rate", f"{summary['review_rate']}%"],
            ["Critical Findings", risk["Critical"]],
            ["High-Risk Findings", risk["High"]],
            [
                "Relationship Indicators",
                summary["relationship_indicator_findings"],
            ],
        ]
        summary_table = Table(
            [["Key Metric", "Result"]] + summary_rows,
            colWidths=[92 * mm, 50 * mm],
            repeatRows=1,
        )
        summary_table.setStyle(
            ReportService._pdf_table_style()
        )
        story.extend([summary_table, Spacer(1, 4 * mm)])

        story.append(
            Paragraph(
                "3. Risk Assessment Summary",
                section_style,
            )
        )
        risk_table = Table(
            [
                ["Low", "Medium", "High", "Critical"],
                [
                    risk["Low"],
                    risk["Medium"],
                    risk["High"],
                    risk["Critical"],
                ],
            ],
            colWidths=[39 * mm] * 4,
        )
        risk_table.setStyle(
            ReportService._pdf_table_style()
        )
        story.extend([risk_table, Spacer(1, 4 * mm)])

        story.append(
            Paragraph(
                "4. Key Audit Observations",
                section_style,
            )
        )
        observations = report.get("audit_observations") or []

        if observations:
            observation_rows = [["Observation / Audit Check", "Occurrences"]]
            for item in observations:
                observation_rows.append(
                    [
                        Paragraph(
                            escape(str(item.get("observation") or "-")),
                            body_style,
                        ),
                        item.get("count") or 0,
                    ]
                )
            observation_table = Table(
                observation_rows,
                colWidths=[126 * mm, 30 * mm],
                repeatRows=1,
            )
            observation_table.setStyle(
                ReportService._pdf_table_style()
            )
            story.append(observation_table)
        else:
            story.append(
                Paragraph(
                    "No configured audit check was triggered in the selected scope.",
                    body_style,
                )
            )

        story.append(Spacer(1, 4 * mm))
        story.append(
            Paragraph(
                "5. Recommended Corrective Actions",
                section_style,
            )
        )
        for index, recommendation in enumerate(
            report.get("recommendations") or [],
            start=1,
        ):
            story.append(
                Paragraph(
                    f"{index}. {escape(str(recommendation))}",
                    body_style,
                )
            )
            story.append(Spacer(1, 1.2 * mm))

        findings = report.get("findings") or []
        if findings:
            story.append(PageBreak())
            story.append(
                Paragraph(
                    "6. Detailed Findings",
                    section_style,
                )
            )

            finding_rows: list[list[Any]] = [
                [
                    "Transaction",
                    "Risk",
                    "Score",
                    "Audit Observation",
                    "Supporting Reason",
                ]
            ]

            for finding in findings:
                finding_rows.append(
                    [
                        str(
                            finding.get("transaction_id")
                            or "-"
                        ),
                        str(
                            finding.get("risk_level")
                            or "-"
                        ),
                        finding.get("risk_score") or 0,
                        Paragraph(
                            escape(
                                str(
                                    finding.get(
                                        "primary_observation"
                                    )
                                    or "-"
                                )
                            ),
                            small_style,
                        ),
                        Paragraph(
                            escape(
                                ReportService._first_reason(
                                    finding.get("reasons")
                                )
                                or "-"
                            ),
                            small_style,
                        ),
                    ]
                )

            findings_table = Table(
                finding_rows,
                colWidths=[
                    25 * mm,
                    18 * mm,
                    16 * mm,
                    49 * mm,
                    50 * mm,
                ],
                repeatRows=1,
            )
            findings_table.setStyle(
                ReportService._pdf_table_style(
                    font_size=7
                )
            )
            story.append(findings_table)

            story.append(PageBreak())
            story.append(
                Paragraph(
                    "7. Supporting Evidence / Technical Appendix",
                    section_style,
                )
            )
            story.append(
                Paragraph(
                    (
                        "This appendix preserves the analytical traceability "
                        "behind each finding. It is intended for auditor review "
                        "and working-paper support rather than management-facing "
                        "interpretation."
                    ),
                    body_style,
                )
            )
            story.append(Spacer(1, 3 * mm))

            evidence_rows: list[list[Any]] = [
                [
                    "Transaction",
                    "Rule Score",
                    "Behaviour Score",
                    "Relationship Score",
                    "Triggered Audit Checks",
                ]
            ]
            for finding in findings:
                technical = (
                    finding.get("technical_evidence")
                    or {}
                )
                evidence_rows.append(
                    [
                        str(
                            finding.get("transaction_id")
                            or "-"
                        ),
                        technical.get("rule_score") or 0,
                        technical.get("anomaly_score") or 0,
                        technical.get("network_score") or 0,
                        Paragraph(
                            escape(
                                ", ".join(
                                    finding.get("triggered_rules")
                                    or []
                                )
                                or "-"
                            ),
                            small_style,
                        ),
                    ]
                )

            evidence_table = Table(
                evidence_rows,
                colWidths=[
                    28 * mm,
                    24 * mm,
                    28 * mm,
                    31 * mm,
                    47 * mm,
                ],
                repeatRows=1,
            )
            evidence_table.setStyle(
                ReportService._pdf_table_style(
                    font_size=6.8
                )
            )
            story.append(evidence_table)

        story.append(Spacer(1, 5 * mm))
        story.append(
            Paragraph(
                "8. Methodology and Limitations",
                section_style,
            )
        )
        story.append(
            Paragraph(
                escape(report.get("methodology_note") or ""),
                body_style,
            )
        )
        story.append(Spacer(1, 2 * mm))
        story.append(
            Paragraph(
                escape(report.get("standards_note") or ""),
                body_style,
            )
        )

        document.build(story)
        return buffer.getvalue()

    @staticmethod
    def _to_xlsx(report: dict[str, Any]) -> bytes:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Executive Summary"

        purple = "5B2C91"
        light_purple = "EEE7F7"
        white = "FFFFFF"
        border_colour = "D8DDE8"

        header_fill = PatternFill(
            "solid",
            fgColor=purple,
        )
        section_fill = PatternFill(
            "solid",
            fgColor=light_purple,
        )
        header_font = Font(
            color=white,
            bold=True,
        )
        section_font = Font(
            color=purple,
            bold=True,
        )
        thin = Side(
            style="thin",
            color=border_colour,
        )

        summary_sheet.append(
            ["INTERNAL AUDIT ANALYTICS REPORT"]
        )
        summary_sheet["A1"].font = Font(
            bold=True,
            size=16,
            color=purple,
        )
        summary_sheet.append(
            ["Status", report.get("report_status")]
        )
        summary_sheet.append(
            [
                "Generated",
                ReportService._format_date(
                    report.get("generated_at")
                ),
            ]
        )
        summary_sheet.append(
            [
                "Client",
                ReportService._primary_client_name(
                    report
                ),
            ]
        )
        summary_sheet.append(
            [
                "Dataset / Scope",
                ReportService._primary_dataset_name(
                    report
                ),
            ]
        )
        summary_sheet.append([])
        summary_sheet.append(
            ["Executive Summary"]
        )
        summary_sheet["A7"].fill = section_fill
        summary_sheet["A7"].font = section_font
        summary_sheet.append(
            [report.get("executive_summary")]
        )
        summary_sheet.append([])
        summary_sheet.append(
            ["Key Metric", "Result"]
        )

        summary = report["summary"]
        risk = summary["risk_counts"]

        for row in [
            [
                "Transactions Reviewed",
                summary["total_transactions"],
            ],
            [
                "Transactions Requiring Review",
                summary["transactions_requiring_review"],
            ],
            [
                "Review Rate",
                summary["review_rate"] / 100
                if summary["review_rate"]
                else 0,
            ],
            ["Low Findings", risk["Low"]],
            ["Medium Findings", risk["Medium"]],
            ["High Findings", risk["High"]],
            ["Critical Findings", risk["Critical"]],
        ]:
            summary_sheet.append(row)

        summary_sheet["B12"].number_format = "0.00%"

        observations_sheet = workbook.create_sheet(
            "Audit Observations"
        )
        observations_sheet.append(
            ["Observation / Audit Check", "Occurrences"]
        )
        for item in report.get("audit_observations") or []:
            observations_sheet.append(
                [
                    item.get("observation"),
                    item.get("count"),
                ]
            )

        findings_sheet = workbook.create_sheet(
            "Detailed Findings"
        )
        finding_headers = [
            "Transaction ID",
            "Client",
            "Dataset",
            "Risk Level",
            "Risk Score",
            "Primary Observation",
            "Supporting Reasons",
            "Triggered Audit Checks",
        ]
        findings_sheet.append(finding_headers)

        for finding in report.get("findings") or []:
            findings_sheet.append(
                [
                    finding.get("transaction_id"),
                    finding.get("client_name"),
                    finding.get("dataset_name"),
                    finding.get("risk_level"),
                    finding.get("risk_score"),
                    finding.get("primary_observation"),
                    " | ".join(
                        finding.get("reasons") or []
                    ),
                    " | ".join(
                        finding.get("triggered_rules")
                        or []
                    ),
                ]
            )

        recommendations_sheet = workbook.create_sheet(
            "Corrective Actions"
        )
        recommendations_sheet.append(
            ["Recommended Corrective Action"]
        )
        for item in report.get("recommendations") or []:
            recommendations_sheet.append([item])

        evidence_sheet = workbook.create_sheet(
            "Technical Evidence"
        )
        evidence_sheet.append(
            [
                "Transaction ID",
                "Rule Score",
                "Behaviour Score",
                "Relationship Score",
                "Technical Sources",
            ]
        )
        for finding in report.get("findings") or []:
            technical = (
                finding.get("technical_evidence")
                or {}
            )
            evidence_sheet.append(
                [
                    finding.get("transaction_id"),
                    technical.get("rule_score"),
                    technical.get("anomaly_score"),
                    technical.get("network_score"),
                    " | ".join(
                        technical.get(
                            "detection_sources"
                        )
                        or []
                    ),
                ]
            )

        methodology_sheet = workbook.create_sheet(
            "Methodology"
        )
        methodology_sheet.append(
            ["Methodology / Limitation"]
        )
        methodology_sheet.append(
            [report.get("methodology_note")]
        )
        methodology_sheet.append(
            [report.get("standards_note")]
        )

        for sheet in workbook.worksheets:
            if sheet.max_row >= 1:
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )

            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = Border(
                        bottom=thin,
                    )
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

            for column_cells in sheet.columns:
                values = [
                    str(cell.value or "")
                    for cell in column_cells
                ]
                max_length = min(
                    max(
                        (
                            len(value)
                            for value in values
                        ),
                        default=10,
                    )
                    + 2,
                    55,
                )
                sheet.column_dimensions[
                    get_column_letter(
                        column_cells[0].column
                    )
                ].width = max_length

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _to_csv(report: dict[str, Any]) -> bytes:
        output = StringIO()
        writer = csv.writer(output)

        writer.writerow(
            [
                "transaction_id",
                "client_name",
                "dataset_name",
                "risk_level",
                "risk_score",
                "primary_observation",
                "supporting_reasons",
                "triggered_audit_checks",
            ]
        )

        for finding in report.get("findings") or []:
            writer.writerow(
                [
                    finding.get("transaction_id"),
                    finding.get("client_name"),
                    finding.get("dataset_name"),
                    finding.get("risk_level"),
                    finding.get("risk_score"),
                    finding.get("primary_observation"),
                    " | ".join(
                        finding.get("reasons") or []
                    ),
                    " | ".join(
                        finding.get("triggered_rules")
                        or []
                    ),
                ]
            )

        return output.getvalue().encode(
            "utf-8-sig"
        )

    @staticmethod
    def _pdf_table_style(
        font_size: int = 8,
    ) -> TableStyle:
        return TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#5B2C91"),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    font_size,
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor("#D8DDE8"),
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#F8FAFC"),
                    ],
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )

    @staticmethod
    def _pdf_detail_table_style() -> TableStyle:
        return TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor("#F1ECFA"),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (0, -1),
                    colors.HexColor("#3B1D76"),
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (0, -1),
                    "Helvetica-Bold",
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor("#D8DDE8"),
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
            ]
        )

    @staticmethod
    def _build_recommendations(
        risk_counts: Counter[str],
        audit_observations: list[dict[str, Any]],
        relationship_indicator_count: int,
        unusual_behaviour_count: int,
    ) -> list[str]:
        recommendations: list[str] = []

        critical = risk_counts.get("Critical", 0)
        high = risk_counts.get("High", 0)

        if critical:
            recommendations.append(
                (
                    f"Perform immediate substantive review of the "
                    f"{critical} Critical finding(s), validate supporting "
                    "documents and record the auditor's conclusion."
                )
            )

        if high:
            recommendations.append(
                (
                    f"Prioritise the {high} High-risk finding(s) for "
                    "evidence verification before finalising the audit review."
                )
            )

        if audit_observations:
            top = audit_observations[0]
            recommendations.append(
                (
                    f"Review the recurring observation '{top['observation']}' "
                    f"({top['count']} occurrence(s)) and assess whether the "
                    "underlying process or control requires corrective action."
                )
            )

        if relationship_indicator_count:
            recommendations.append(
                (
                    f"Validate vendor master information and supporting "
                    f"documents for {relationship_indicator_count} finding(s) "
                    "with relationship indicators such as shared identifiers."
                )
            )

        if unusual_behaviour_count:
            recommendations.append(
                (
                    f"Perform additional substantive procedures on "
                    f"{unusual_behaviour_count} transaction(s) showing unusual "
                    "behaviour and document whether the deviations have a "
                    "valid business explanation."
                )
            )

        if not recommendations:
            recommendations.append(
                (
                    "No significant exception requiring a specific corrective "
                    "action was identified in the selected scope. Retain the "
                    "analysis output as supporting audit documentation."
                )
            )

        return recommendations

    @staticmethod
    def _build_executive_summary(
        total_transactions: int,
        total_findings: int,
        high_critical: int,
        critical: int,
        selected_clients: list[dict[str, Any]],
        selected_datasets: list[dict[str, Any]],
    ) -> str:
        client_text = (
            str(
                selected_clients[0].get(
                    "client_name"
                )
            )
            if len(selected_clients) == 1
            else f"{len(selected_clients)} clients"
        )
        dataset_text = (
            str(
                selected_datasets[0].get(
                    "dataset_name"
                )
            )
            if len(selected_datasets) == 1
            else f"{len(selected_datasets)} datasets"
        )

        return (
            f"The selected audit scope for {client_text} covered "
            f"{total_transactions} transaction(s) across {dataset_text}. "
            f"The analytics review identified {total_findings} transaction(s) "
            f"requiring auditor review, including {high_critical} High/Critical "
            f"finding(s) and {critical} Critical finding(s). The results should "
            "be evaluated together with source documents, management explanations "
            "and the auditor's professional judgement before final conclusions "
            "or corrective actions are issued."
        )

    @staticmethod
    def _scope_text(
        report_type: str,
        selected_clients: list[dict[str, Any]],
        selected_datasets: list[dict[str, Any]],
        selected_analyses: list[dict[str, Any]],
    ) -> str:
        client_names = [
            str(item.get("client_name") or "")
            for item in selected_clients
            if item.get("client_name")
        ]
        dataset_names = [
            str(item.get("dataset_name") or "")
            for item in selected_datasets
            if item.get("dataset_name")
        ]

        return (
            f"Report type: {report_type.title()}. "
            f"Client scope: {', '.join(client_names) if client_names else 'All selected clients'}. "
            f"Dataset scope: {', '.join(dataset_names) if dataset_names else 'All selected datasets'}. "
            f"Analysis runs included: {len(selected_analyses)}."
        )

    @staticmethod
    def _report_title(
        report_type: str,
    ) -> str:
        titles = {
            "executive": (
                "Internal Audit Analytics - Executive Report"
            ),
            "client": (
                "Internal Audit Analytics - Client Report"
            ),
            "dataset": (
                "Internal Audit Analytics - Dataset Review"
            ),
            "analysis": (
                "Internal Audit Analytics - Analysis Report"
            ),
            "investigation": (
                "Internal Audit Analytics - Investigation Report"
            ),
        }
        return titles[report_type]

    @staticmethod
    def _as_string_list(
        value: Any,
    ) -> list[str]:
        if value is None:
            return []

        if isinstance(value, list):
            return [
                str(item).strip()
                for item in value
                if str(item).strip()
            ]

        if isinstance(value, str):
            stripped = value.strip()

            if not stripped:
                return []

            try:
                parsed = json.loads(stripped)

                if isinstance(parsed, list):
                    return [
                        str(item).strip()
                        for item in parsed
                        if str(item).strip()
                    ]
            except json.JSONDecodeError:
                pass

            return [
                item.strip()
                for item in stripped.split("|")
                if item.strip()
            ]

        return [str(value).strip()]

    @staticmethod
    def _reason_list(
        value: Any,
    ) -> list[str]:
        return ReportService._as_string_list(value)

    @staticmethod
    def _network_reasons(
        finding: dict[str, Any],
    ) -> list[str]:
        explanation = finding.get("explanation")

        if isinstance(explanation, dict):
            return ReportService._as_string_list(
                explanation.get(
                    "network_reasons"
                )
            )

        if isinstance(explanation, str):
            try:
                parsed = json.loads(explanation)

                if isinstance(parsed, dict):
                    return ReportService._as_string_list(
                        parsed.get(
                            "network_reasons"
                        )
                    )
            except json.JSONDecodeError:
                pass

        return []

    @staticmethod
    def _number(value: Any) -> float:
        try:
            return round(
                float(value or 0),
                4,
            )
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _risk_rank(
        risk_level: str,
    ) -> int:
        return {
            "Critical": 4,
            "High": 3,
            "Medium": 2,
            "Low": 1,
        }.get(risk_level, 0)

    @staticmethod
    def _first_reason(
        value: Any,
    ) -> str:
        if isinstance(value, list):
            return (
                str(value[0]).strip()
                if value
                else ""
            )

        if not value:
            return ""

        return str(value).split("|")[0].strip()

    @staticmethod
    def _format_date(
        value: Any,
    ) -> str:
        if not value:
            return "-"

        return (
            str(value)
            .replace("T", " ")
            .replace("Z", " UTC")
        )

    @staticmethod
    def _primary_client_name(
        report: dict[str, Any],
    ) -> str:
        clients = report.get("clients") or []

        if len(clients) == 1:
            return str(
                clients[0].get("client_name")
                or "-"
            )

        return (
            f"{len(clients)} clients"
            if clients
            else "All selected clients"
        )

    @staticmethod
    def _primary_dataset_name(
        report: dict[str, Any],
    ) -> str:
        datasets = report.get("datasets") or []

        if len(datasets) == 1:
            return str(
                datasets[0].get("dataset_name")
                or "-"
            )

        return (
            f"{len(datasets)} datasets"
            if datasets
            else "All selected datasets"
        )
